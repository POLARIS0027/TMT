# 納品作業関連

import os
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from config import load_config

class DeliveryHelper:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.config, _ = load_config()  # 설정 파일 불러오기
        self.sheet_name = self.config.get("sheet_name", "試験表")  # 기본값 fallback

    def set_font_to_meiryo(self):
        for root, _, files in os.walk(self.folder_path):
            for file in files:
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file)
                    st.write(f"フォントを変更中...: {file_path}")
                    try:
                        workbook = load_workbook(file_path)
                        for sheet in workbook.worksheets:
                            for row in sheet.iter_rows():
                                for cell in row:
                                    cell.font = Font(name='メイリオ')
                        workbook.save(file_path)
                        workbook.close()
                        st.write("Done")
                    except Exception as e:
                        st.error(f"Error setting font in {file_path}: {e}")

    def align_cells_left_top(self):
        for root, _, files in os.walk(self.folder_path):
            for file in files:
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file)
                    st.write(f"左上に整列中...: {file_path}")
                    try:
                        workbook = load_workbook(file_path)
                        if self.sheet_name in workbook.sheetnames:
                            sheet = workbook[self.sheet_name]
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value is not None:
                                        cell.alignment = Alignment(horizontal='left', vertical='top')
                            st.write(f"'{self.sheet_name}' シート整列完了: {file_path}")
                        else:
                            st.write(f"'{self.sheet_name}' シートが存在しません: {file_path}")
                        workbook.save(file_path)
                        workbook.close()
                    except Exception as e:
                        st.error(f"Error aligning cells in {file_path}: {e}")

    def fill_blank_cells_in_range(self):
        for root, _, files in os.walk(self.folder_path):
            for file in files:
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file)
                    st.write(f"処理中: {file_path}")
                    try:
                        workbook = load_workbook(file_path)
                        if self.sheet_name in workbook.sheetnames:
                            sheet = workbook[self.sheet_name]
                            data = sheet.values
                            rows = list(data)
                            df = pd.DataFrame(rows).dropna(how='all', axis=0).dropna(how='all', axis=1)
                            df = df.fillna('-')
                            for row_idx, row in enumerate(df.values, start=1):
                                for col_idx, value in enumerate(row, start=1):
                                    sheet.cell(row=row_idx, column=col_idx, value=value)
                            workbook.save(file_path)
                            workbook.close()
                            st.write(f"'{self.sheet_name}' シート空白処理完了: {file_path}")
                        else:
                            st.write(f"'{self.sheet_name}' シートが存在しません: {file_path}")
                    except Exception as e:
                        st.error(f"Error processing {file_path}: {e}")

    def set_zoom_to_100(self):
        for root, _, files in os.walk(self.folder_path):
            for file in files:
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file)
                    st.write(f"拡大比率を調整中: {file_path}")
                    try:
                        workbook = load_workbook(file_path)
                        for sheet in workbook.worksheets:
                            sheet.sheet_view.zoomScale = 100
                        workbook.save(file_path)
                        workbook.close()
                        st.write("Done")
                    except Exception as e:
                        st.error(f"Error setting zoom in {file_path}: {e}")
